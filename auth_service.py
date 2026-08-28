"""
Login server-side untuk INTRA.

PENTING soal keamanan (supaya jelas kenapa desainnya begini): file
users.xlsx dibaca di SINI, di server -- TIDAK PERNAH dikirim ke browser
dalam bentuk apapun. Jadi user (via Inspect Element / tab Network) tidak
akan pernah bisa melihat isi users.xlsx sama sekali, terlepas dari apakah
file itu sendiri "dienkripsi" atau tidak -- bedanya jauh dari kalau ini
app React/JS yang jalan di browser (di situ SEMUA data yang dipakai buat
cek login otomatis ikut terkirim ke browser, makanya rawan).

Password tetap di-hash (werkzeug generate_password_hash/check_password_hash
-- pakai scrypt, sudah termasuk salt otomatis) sebagai lapisan tambahan,
supaya kalau file users.xlsx ini somehow bocor/ke-commit ke git dsb,
password asli user tetap tidak langsung kebaca.
"""
import functools
import os

import openpyxl
from flask import session, redirect, url_for, request
from werkzeug.security import check_password_hash

USERS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "users.xlsx")

ROLE_LABELS = {
    "developer": "Developer",
    "admin": "Admin",
    "manager": "Manager",
    "user": "User",
}


def load_users():
    """Baca users.xlsx tiap kali dipanggil -- SENGAJA tidak di-cache, supaya
    perubahan (tambah/hapus/reset password user) di file langsung kepakai
    tanpa perlu restart server. File-nya kecil (jumlah user terbatas), jadi
    baca ulang tiap request tidak masalah dari sisi performa."""
    if not os.path.exists(USERS_FILE):
        return []
    wb = openpyxl.load_workbook(USERS_FILE, read_only=True, data_only=True)
    ws = wb["users"] if "users" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    users = []
    for r in rows:
        if not r or not r[0]:
            continue
        nik, name, password_hash, role, project = (list(r) + [None] * 5)[:5]
        users.append({
            "nik": str(nik).strip(),
            "name": (name or "").strip(),
            "password_hash": (password_hash or "").strip(),
            "role": (role or "user").strip().lower(),
            "project": (project or "").strip().upper() or "ALL",
        })
    return users


def verify_login(password):
    """Return dict user kalau password cocok dengan salah satu user di
    users.xlsx, None kalau tidak ada yang cocok.

    NIK sudah tidak diminta lagi di form login -- password sendiri yang
    jadi kredensial buat kenalin user-nya (makanya tiap user WAJIB punya
    password unik masing-masing di users.xlsx, jangan sampai 2 user pakai
    password yang sama, nanti yang kepilih cuma yang baris pertama
    ketemu). Role & akses per-project (kolom role/project) tetap jalan
    seperti biasa karena hasilnya tetap dict user yang lengkap.

    Tiap baris di-bungkus try/except: kalau ADA SATU user yang
    password_hash-nya rusak/format-nya salah di users.xlsx (mis. ke-korup
    waktu di-paste ke Excel), baris itu di-skip aja -- supaya tidak bikin
    SEMUA orang gagal login gara-gara satu baris yang rusak.

    Catatan performa: check_password_hash (scrypt) sengaja lambat demi
    keamanan, dan di sini di-loop ke semua user tiap kali login. Untuk
    jumlah user yang kecil (internal tool) ini masih aman; kalau jumlah
    user sudah banyak (puluhan+), pertimbangkan balikin opsi isi
    NIK/username lagi supaya lookup-nya O(1) per user, bukan di-scan
    semua."""
    if not password:
        return None
    for u in load_users():
        stored_hash = u["password_hash"]
        if not stored_hash:
            continue
        try:
            if check_password_hash(stored_hash, password):
                return u
        except ValueError:
            # Format hash user ini rusak (mis. field password_hash di
            # users.xlsx ke-korup/terpotong). Skip baris ini saja.
            continue
    return None


def current_user():
    return session.get("user")


def can_access_menu(user, key):
    """Sama aturan dengan draft React sebelumnya: developer/admin/manager
    bebas akses semua menu, role 'user' cuma menu sesuai kolom project-nya
    di users.xlsx ('ALL' juga bebas akses semua)."""
    if not user:
        return False
    if user["role"] in ("developer", "admin", "manager"):
        return True
    return user["project"] == "ALL" or user["project"] == key


def login_required(view_func):
    """Decorator: redirect ke /login kalau belum login. Simpan halaman yang
    dituju di ?next= supaya begitu login sukses, langsung diarahkan balik
    ke situ (bukan selalu ke halaman pilihan project)."""
    @functools.wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get("user"):
            return redirect(url_for("login", next=request.path))
        return view_func(*args, **kwargs)
    return wrapped